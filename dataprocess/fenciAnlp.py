# coding:utf-8
# 分词以及SnowNLP情感分析，对句子评分及对分词结果评分
import xlrd
from snownlp import SnowNLP
import re
import emoji
import jieba
from jieba import posseg
import openpyxl
# 在这里遇到的问题：excel文档用的是utf-8,控制台用的是gbk输出，最好自己去更改一下设置就好
# （在file-settings-fileEncodings-utf-8）两个都选成utf-8，这样就可以直接输出

# 保存分词结果
workbook1 = openpyxl.Workbook()
worksheet1 = workbook1.active
worksheet1.title = "南京景点全部评论分词"

xlsx_path = "..\mafengwo\\spiders\\Nanjing.xlsx"
workbook = xlrd.open_workbook(xlsx_path, 'utf-8')
sheet1 = workbook.sheet_by_index(0)  # 第一个工作簿
print(sheet1.nrows)


def stripword(sequence):      # 剔除停用词
    wordlist = []
    # 获取停用词表
    stop = open('SCUstopwprd.txt', 'r+', encoding='utf-8')
    stopword = stop.read().split("\n")
    # 遍历分词表
    for key in sequence.split('/'):
        # 去除停用词，去除单字，去除重复词  不要去除单字了
        if not(key.strip() in stopword):
            wordlist.append(key)
    stop.close()
    return '/'.join(wordlist)


def data_process_jb():     # 结巴分词
    for i in range(1, sheet1.nrows):  # (173, 174)
        # 去除emoji表情，emojize()：根据code生成emoji表情，demojize()：将emoji表情解码为code#
        # []内放需要处理掉的数据 [^\u4e00-\u9fa5]+去除非中文
        review = re.sub(r'[\',\\xa0]', '', sheet1.cell(i, 2).value)
        review_text = re.sub('(:.+?:)', '', emoji.demojize(review))
        # print(review_text)   # 打印评论
        snow = SnowNLP(review_text)
        review_text = snow.han        # 繁转简，review_text是文本处理后的最终结果
        analysis = "/".join(jieba.cut(review_text))
        sentence = stripword(analysis)  # sentence是分词之后的结果，有/号
        # print(sentence)
        # print(SnowNLP(review_text).sentiments)
        worksheet1.cell(i, 1, sentence)  # 保存分词结果
        worksheet1.cell(i, 2, float(SnowNLP(review_text).sentiments))  # 保存整句话情感评分
        analysis_words = [(word.word, word.flag) for word in posseg.cut(sentence)]
        #print(analysis_words)
        # 只提取标签为a、d、v的三类词
        keywords = [x for x in analysis_words if x[1] in ['a', 'd', 'v']]
        #  print(keywords)
        # 根据关键词的标签提取出关键字以后，这个时候可以将情感标记去除只保留关键字就可以了。
        keywords = [x[0] for x in keywords]
        snownlp_process(keywords)
        # print("----------------")
    workbook1.save(filename='FenciNJ.xlsx')


# SnowNLP分析出结果
def snownlp_process(keywords):
    pos_num = 0         # 正面情绪数
    neg_num = 0         # 负面情绪数
    # 循环遍历关键字列表中的每个单词
    for word in keywords:
        sl = SnowNLP(word)
        # 检查单词的情感是否大于0.5。
        if sl.sentiments > 0.5:
            pos_num = pos_num + 1
        else:
            neg_num = neg_num + 1
        # print(word, str(sl.sentiments))
        with open("评论分词情感分析" + r'.txt', 'a+', encoding='utf-8') as f:   # 写入txt文件
            f.write(word +" " + str(sl.sentiments) + "\n")


data_process_jb()










# str2 = "人太多了"
# s1 = SnowNLP(str2)
# print(s1.words)
# print(s1.sentiments)
# print("--------")
# s2 = jieba.lcut(str2)
# print(s2)
# print(s2.sentiments)

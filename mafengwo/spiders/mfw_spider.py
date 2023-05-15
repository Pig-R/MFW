# coding:utf-8
import scrapy
from ..items import MafengwoItem
import openpyxl
from scrapy import Request, Spider
import time
import xlrd
import xlutils.copy
# from scrap_mfw.items import ScrapMfwItem #这样写会有红色波浪线
import requests
from requests.utils import add_dict_to_cookiejar
from requests.packages.urllib3.exceptions import InsecureRequestWarning
# 关闭ssl验证提示


class MfwSpiderSpider(scrapy.Spider):
    name = 'mfw_spider'
    allowed_domains = ['m.mafengwo.cn']
    start_urls = ['https://m.mafengwo.cn/']
    # 总统府https://m.mafengwo.cn/poi/comment_5429143.html
    # 景点id
    place = ['5429143', '3452', '1529', '3601', '2293952', '4789', '3461', '3462']
    place_url = []  # 放拼好的url
    # 全部评论
    workbook1 = openpyxl.Workbook()  # 创建一个Workbook对象
    worksheet1 = workbook1.active
    worksheet1.title = "南京景点"  # excel下面的名字
    Project = ['景点名称', '景点星级', '景点评论']
    # 好评
    workbook2 = openpyxl.Workbook()
    worksheet2 = workbook2.active
    worksheet2.title = "南京景点好评"
    # 差评
    workbook3 = openpyxl.Workbook()
    worksheet3 = workbook3.active
    worksheet3.title = "南京景点差评"
    i = 2  # 全部评论
    j = 2  # 好评
    k = 2  # 差评

    def start_requests(self):
        for j in range(len(self.Project)):
            self.worksheet1.cell(1, j + 1, self.Project[j])
            self.worksheet2.cell(1, j + 1, self.Project[j])
            self.worksheet3.cell(1, j + 1, self.Project[j])
        for p in self.place:
            self.place_url.append(self.start_urls[0]+"poi/comment_" + p + ".html")
        # 循环抓取全部评论、好评、差评
        for url in self.place_url:
            yield scrapy.Request(url=url, callback=self.parse, dont_filter=True,
                                 meta={'url': url, 'status': 0})
        time.sleep(1)
        for url in self.place_url:
            yield scrapy.Request(url=url, callback=self.parse, dont_filter=True,
                                 meta={'url': url, 'status': 1})
        time.sleep(1)
        for url in self.place_url:
            yield scrapy.Request(url=url, callback=self.parse, dont_filter=True,
                                 meta={'url': url, 'status': 2})  # 在这去中间件
        # yield scrapy.Request(url=self.place_url[0], callback=self.parse, dont_filter=True,
        #                      meta={'url': self.place_url[0], 'status': self.status})  # 在这去中间件

    def parse(self, response):
        flag = response.meta.get('status')
        title = response.xpath("/html/head/title/text()").extract()
        title = str(title).strip(']').strip('[').strip("'").replace("评论 - 马蜂窝", '')
        percent = response.xpath("//*[@class='wrapper']/section/div/div/div[2]/p/text()").extract()
        percent = str(percent).strip(']').strip('[').strip("'").replace('的蜂蜂推荐', '')
        self.worksheet1.cell(self.i, 1, title)
        self.worksheet1.cell(self.i, 2, percent)

        self.worksheet2.cell(self.j, 1, title)
        self.worksheet2.cell(self.j, 2, percent)

        self.worksheet3.cell(self.k, 1, title)
        self.worksheet3.cell(self.k, 2, percent)
        ReviewList = response.xpath("//*[@class='wrapper']/section/ul[@class='list']/li")
        for reviews in ReviewList:
            review = reviews.xpath(".//div[2]/text()").extract()
            r = ''.join(str(review).strip("[").strip("]").strip("'").replace("\\n", '').\
                replace("\\u3000", '').replace("\xa0", '').split())
            if flag == 0:  # 全部评论
                self.worksheet1.cell(self.i, 3, r)
                self.i += 1
            if flag == 1:  # 好评
                self.worksheet2.cell(self.j, 3, r)
                self.j += 1
            if flag == 2:
                self.worksheet3.cell(self.k, 3, r)
                self.k += 1
        self.workbook1.save(filename='Nanjing.xlsx')
        self.workbook2.save(filename='NanjingHao.xlsx')
        self.workbook3.save(filename='NanjingHuai.xlsx')

























# 之前之前的代码

 # def parse(self, response):
 #        # url = response.meta.get('url')
 #        # item = MafengwoItem()
 #        title = response.xpath("/html/head/title/text()").extract()
 #        # item['Title'] = title
 #        title = str(title).strip(']').strip('[').strip("'").replace("评论 - 马蜂窝", '')
 #        percent = response.xpath("//*[@class='wrapper']/section/div/div/div[2]/p/text()").extract()
 #        # item['Percent'] = percent
 #        percent = str(percent).strip(']').strip('[').strip("'").replace('的蜂蜂推荐', '')
 #        self.worksheet1.cell(self.i, 1, title); self.worksheet2.cell(self.i, 1, title);
 #        self.worksheet3.cell(self.i, 1, title)
 #        self.worksheet1.cell(self.i, 2, percent);self.worksheet2.cell(self.i, 2, percent)
 #        self.worksheet3.cell(self.i, 2, percent)
 #        restr = ""
 #        ReviewList = response.xpath("//*[@class='wrapper']/section/ul[@class='list']/li")
 #        for reviews in ReviewList:
 #            review = reviews.xpath(".//div[2]/text()").extract()
 #            r = ''.join(str(review).strip("[").strip("]").strip("'").replace("\\n", '').\
 #                replace("\\u3000", '').replace("\xa0", '').split())
 #            restr = restr + ''.join(str(review).replace("\\n", '').replace("\\u3000", '').replace("\xa0", '').split())
 #            self.worksheet1.cell(self.i, 3, r)
 #            self.i += 1
 #            # item['Reviews'] = restr
 #        self.status = 1
 #        self.workbook1.save(filename='Nanjing.xlsx')
 #        self.workbook2.save(filename='NanjingHao.xlsx')
 #        self.workbook3.save(filename='NanjingHuai.xlsx')

# if self.status == 1:
        #     yield scrapy.Request(url, callback=self.goodreviews, dont_filter=True,
        #                          meta={'status': self.status, 'url': url})
        # self.status = 2
        # if self.status == 2:
        #     yield scrapy.Request(url, callback=self.badreviews, dont_filter=True,
        #                          meta={'status': self.status, 'url': url})
        # self.status = 0
        # if self.status == 0:
        #     self.u += 1
        #     yield scrapy.Request(self.place_url[self.u], callback=self.parse, dont_filter=True,
        #                          meta={'status': self.status, 'url': self.place_url[self.u]})
       # yield item
    # 上面是先按顺序执行再跳回到yield 的内容